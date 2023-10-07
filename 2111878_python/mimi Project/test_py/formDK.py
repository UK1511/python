from openpyxl import *
from tkinter import *

wb = load_workbook('C:\\Users\\CICT\\Desktop\\test_py\\form.xlsx')
sheet = wb.active

def execl():
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 60
    sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 100
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 10
    sheet.column_dimensions['G'].width = 15
    sheet.column_dimensions['H'].width = 15
    
    
    sheet.cell(row=1, column=1).value = 'Mã số sinh viên'
    sheet.cell(row=1, column=2).value = 'Họ và Tên'
    sheet.cell(row=1, column=3).value = 'Ngày sinh'
    sheet.cell(row=1, column=4).value = 'Email'
    sheet.cell(row=1, column=5).value = 'Số điện thoại'
    sheet.cell(row=1, column=6).value = 'Học kỳ'
    sheet.cell(row=1, column=7).value = 'Năm học'
    sheet.cell(row=1, column=8).value = 'Môn học'
    
def focus1(event):
    name_field.focus_set()

def focus2(event):
    dateBirth_field.focus_set()

def focus3(event):
    email_field.focus_set()
    
def focus4(event):
    contact_field.focus_set()
    
def focus5(event):
    sem_field.focus_set()

def focus6(event):
    yearschool_field.focus_set()

def focus7(event):
    subject_field.focus_set()    
    
    
def clear():
    mssv_field.delete(0, END)
    name_field.delete(0, END)
    dateBirth_field.delete(0, END)
    email_field.delete(0, END)
    contact_field.delete(0, END)
    sem_field.delete(0, END)
    yearschool_field.delete(0, END)
    subject_field.delete(0, END)
    
def insert():
    if (mssv_field.get()== "" and
        name_field.get() == "" and
        dateBirth_field.get() == "" and
        email_field.get() == "" and
        contact_field.get() == "" and
        sem_field.get() == "" and
        yearschool_field.get() == "" and
        IntVar.get() == ""):
        print('Empty input')
        
    else:
        current_row = sheet.max_row
        
        sheet.cell(row=current_row + 1, column = 1).value = mssv_field.get()
        sheet.cell(row=current_row + 1, column = 2).value = name_field.get()
        sheet.cell(row=current_row + 1, column = 3).value = dateBirth_field.get()
        sheet.cell(row=current_row + 1, column = 4).value = email_field.get()
        sheet.cell(row=current_row + 1, column = 5).value = contact_field.get()
        sheet.cell(row=current_row + 1, column = 6).value = sem_field.get()
        sheet.cell(row=current_row + 1, column = 7).value = yearschool_field.get()
        sheet.cell(row=current_row + 1, column = 8).value = IntVar.get()
        
        wb.save('C:\\Users\\CICT\\Desktop\\test_py\\form.xlsx')
        
        mssv_field.focus_set()
        clear()
        
if __name__ == '__main__':
    root = Tk()
    root.configure(background='light green')
    
    root.title('Form đăng ký')
    root.geometry('500x300')
    execl()
    
    heading = Label(root, text= 'THÔNG TIN ĐĂNG KÝ HỌC PHẦN', bg= "light green")
    mssv = Label(root, text="Mã số sinh viên" , bg= "light green")
    name = Label(root, text="Họ tên", bg= "light green")
    dateBirth = Label(root, text="Ngày sinh ", bg= "light green")
    email = Label(root, text= "Email", bg="light green")
    sdt = Label(root, text="Số điện thoại ", bg="light green")
    sem = Label(root, text="Học kỳ", bg="light green")
    yearschool = Label(root, text="Năm học", bg="light green")
    subject = Label(root, text="Chọn môn học", bg="light green")
    
    
    heading.grid(row=0, column=1)
    mssv.grid(row=1,column=0)
    name.grid(row=2, column=0)
    dateBirth.grid(row=3, column=0)
    email.grid(row=4,column=0)
    sdt.grid(row=5, column=0)
    sem.grid(row=6, column=0)
    yearschool.grid(row=7, column=0)
    subject.grid(row=8, column=0)
    
    mssv_field = Entry(root)
    name_field = Entry(root)
    dateBirth_field = Entry(root)
    email_field = Entry(root)
    contact_field =Entry(root)
    sem_field =Entry(root)
    yearschool_field =Entry(root)
    subject_field = Entry(root)
    
    mssv_field.bind("<Return>", focus1)
    name_field.bind("<Return>", focus2)
    dateBirth_field.bind("<Return>", focus3)
    email_field.bind("<Return>", focus4)
    contact_field.bind("<Return>", focus5)
    sem_field.bind("<Return>", focus6)
    yearschool_field.bind("<Return>", focus7)
    
    
    mssv_field.grid(row=1, column=1, ipadx="100")
    name_field.grid(row=2, column=1, ipadx="100")
    dateBirth_field.grid(row=3, column=1, ipadx="100")
    email_field.grid(row=4, column=1, ipadx="100")
    contact_field.grid(row=5, column=1, ipadx="100")
    sem_field.grid(row=6, column=1, ipadx="100")
    yearschool_field.grid(row=7, column=1, ipadx="100")
    
    check1 = IntVar()
    Checkbutton(root, text="Lập trình python", background="light green").place(x= 90, y=170)
    check2 = IntVar()
    Checkbutton(root, text="Lập trình java", background="light green").place(x= 250, y= 170)
    check3 = IntVar()
    Checkbutton(root, text="Công nghệ phần mềm",background="light green").place(x=90, y=220)
    check4 = IntVar()
    Checkbutton(root, text="Phát triển ứng dụng web", background="light green").place(x= 250, y= 220)
    execl()

    submit = Button(root, text="Đăng ký", bg="green", command=insert)
    submit.config(width=8, height=1)
    submit.place(x=125, y= 255)
    
    exit = Button(root, text="Thoát", bg="green", command=exit)
    exit.config(width=8, height=1)
    exit.place(x=300, y=255)
    
    root.mainloop()